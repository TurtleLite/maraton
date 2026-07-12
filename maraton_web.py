import io
import os
import traceback
import psycopg2
import psycopg2.extras
from datetime import datetime
from flask import Flask, request, jsonify, render_template_string, send_file, g
import openpyxl

app = Flask(__name__)
app.static_folder = "static"

DB_URL = os.environ.get("DATABASE_URL", "postgresql://maraton_db_yu80_user:PGKNIB3F5HTynSqz7Vx6x01vJcQVG3bD@dpg-d97vq9qabeoc739aqnmg-a.virginia-postgres.render.com/maraton_db_yu80?sslmode=require")

def get_db():
    if 'db' not in g:
        try:
            conn = psycopg2.connect(DB_URL, connect_timeout=10)
            conn.autocommit = True
            g.db = conn
        except Exception as e:
            traceback.print_exc()
            return None
    return g.db

@app.teardown_request
def cerrar_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except:
            pass

def init_db():
    conn = get_db()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS carrera (id SERIAL PRIMARY KEY, iniciada BOOLEAN DEFAULT FALSE, hora_inicio TIMESTAMP)")
        cur.execute("CREATE TABLE IF NOT EXISTS corredores (id SERIAL PRIMARY KEY, dorsal VARCHAR(20) UNIQUE NOT NULL, nombre VARCHAR(200) NOT NULL, tiempo_llegada TIMESTAMP, posicion INTEGER)")
        cur.execute("ALTER TABLE corredores ADD COLUMN IF NOT EXISTS categoria VARCHAR(20) DEFAULT ''")
        cur.execute("ALTER TABLE corredores ADD COLUMN IF NOT EXISTS posicion_categoria INTEGER")
        cur.execute("SELECT COUNT(*) FROM carrera")
        if cur.fetchone()[0] == 0:
            cur.execute("INSERT INTO carrera (iniciada) VALUES (FALSE)")
        cur.close()
    except Exception as e:
        traceback.print_exc()

LOGO_IZQ = "/static/BURROTON 2026.png"
LOGO_DER = "/static/LOGO SAN BENITO JOSE.png"

HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Burrotón San Benito José</title>
<link rel="icon" href="/static/BURROTON 2026.png">
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;900&display=swap" rel="stylesheet">
<style>
html, body { height: 100%; margin: 0; padding: 0; }
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f4f6f9; color: #1c2838; min-height: 100vh; }
::selection { background: #c9953e; color: #fff; }
::-webkit-scrollbar { width: 7px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #c8d4e0; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #a8bcd0; }
.container { width: 100%; min-height: 100vh; margin: 0; padding: 0; background: #f4f6f9; display: flex; flex-direction: column; }
.header-wrap { background: #fff; padding: 28px 44px 20px; position: relative; border-bottom: 2px solid #e8ecf0; }
.header-wrap::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: linear-gradient(90deg, #2c5f8a, #c9953e, #2c5f8a); pointer-events: none; }
.header { display: flex; align-items: center; justify-content: space-between; gap: 24px; position: relative; z-index: 1; }
.header img { height: 120px; width: auto; object-fit: contain; transition: transform .4s cubic-bezier(.34,1.56,.64,1); }
.header img:hover { transform: scale(1.05); }
.header h1 { font-size: 1.9rem; color: #1c2838; text-align: center; flex: 1; font-weight: 800; letter-spacing: -.01em; font-family: 'Inter', -apple-system, sans-serif; animation: fadeIn 1s ease .2s both; line-height: 1.3; }
.main-content { padding: 28px 44px 80px; flex: 1; }
.barra { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 12px; }
.barra button, .barra input, .barra select { padding: 9px 16px; font-size: .85rem; border-radius: 8px; border: 1px solid #dce0e8; outline: none; font-family: inherit; }
.barra input:focus { border-color: #8a9aa8; box-shadow: 0 0 0 2px rgba(44, 62, 106, .08); }
.barra select:focus { border-color: #8a9aa8; box-shadow: 0 0 0 2px rgba(44, 62, 106, .08); }
.barra button { background: #2c5f8a; color: #fff; border: none; cursor: pointer; font-weight: 500; transition: all .2s ease; }
.barra button:hover { background: #1e4a70; box-shadow: 0 4px 12px rgba(44, 62, 106, .15); }
.barra button:active { transform: scale(.97); }
.barra button:disabled { opacity: .3; cursor: not-allowed; box-shadow: none; transform: none; }
.barra button.verde { background: #3a8a5a; }
.barra button.verde:hover { background: #2a7a4a; box-shadow: 0 4px 12px rgba(58, 138, 90, .2); }
.barra button.rojo { background: #b85048; }
.barra button.rojo:hover { background: #a84038; box-shadow: 0 4px 12px rgba(184, 80, 72, .2); }
.barra input { flex: 1; min-width: 120px; background: #fff; transition: all .2s ease; color: #1c2838; }
.barra input:focus { background: #fff; }
.barra select { background: #fff; color: #1c2838; cursor: pointer; transition: border-color .2s ease; }
.seccion { background: #fff; border: 1px solid #eaecf0; border-radius: 10px; padding: 22px 26px; margin-bottom: 18px; box-shadow: 0 1px 3px rgba(0,0,0,.03); transition: all .25s ease; animation: slideUp .45s ease-out both; }
.seccion:nth-child(2) { animation-delay: .08s; }
.seccion:nth-child(3) { animation-delay: .16s; }
.seccion:nth-child(4) { animation-delay: .24s; }
.seccion:hover { box-shadow: 0 4px 16px rgba(44, 62, 106, .06); border-color: #d0d4d8; }
.seccion h2 { font-size: .7rem; color: #8a9aa8; font-weight: 600; text-transform: uppercase; letter-spacing: .1em; margin-bottom: 16px; }
.seccion .fila { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }
.seccion .fila.espaciada { margin-top: 14px; }
#modo-registro { padding: 10px 18px; font-size: .875rem; border-radius: 10px; border: 1px solid #d0dce8; outline: none; font-family: inherit; background: #fff; color: #1e2a3a; cursor: pointer; transition: border-color .2s ease; }
#modo-registro:focus { border-color: #4a7fc8; box-shadow: 0 0 0 3px rgba(74, 127, 200, .2); }
.estado { padding: 8px 18px; border-radius: 10px; font-weight: 600; display: inline-flex; align-items: center; gap: 6px; font-size: .875rem; border: 1px solid; transition: all .3s ease; }
.estado.parada { background: #faf3e0; color: #8a6a30; border-color: #f0d080; }
.estado.andando { background: #e0f4e8; color: #2a7a48; border-color: #80d8a0; }
.buscar-wrap { margin-bottom: 14px; position: relative; animation: slideUp .45s ease-out .28s both; }
.buscar-wrap input { width: 100%; padding: 14px 18px; font-size: .9rem; border-radius: 8px; border: 1px solid #dce0e8; outline: none; font-family: inherit; background: #fff; transition: all .2s ease; color: #1c2838; }
.buscar-wrap input:focus { border-color: #8a9aa8; box-shadow: 0 0 0 2px rgba(44, 62, 106, .08); }
.buscar-wrap input::placeholder { color: #b0bcc8; font-weight: 400; }
table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: .85rem; }
th, td { padding: 11px 14px; text-align: left; border-bottom: 1px solid #e8eef4; }
th { background: #f0f2f5; font-weight: 600; color: #3a4a58; font-size: .75rem; text-transform: uppercase; letter-spacing: .06em; position: sticky; top: 0; z-index: 2; border-bottom: 2px solid #dce0e8; }
tr:last-child td { border-bottom: none; }
tr:last-child td:first-child { border-radius: 0 0 0 8px; }
tr:last-child td:last-child { border-radius: 0 0 8px 0; }
tr td { transition: background .12s ease; }
tr:hover td { background: #eef2f6; }
tbody tr:nth-child(even) td { background: #fafbfc; }
tbody tr:nth-child(even):hover td { background: #eef2f6; }
td button { transition: all .2s ease; cursor: pointer; }
td button:hover { transform: scale(1.1); }
td button:active { transform: scale(.88); }
.tabla-wrap { border-radius: 10px; border: 1px solid #dce0e8; overflow: auto; animation: slideUp .45s ease-out .32s both; background: #fff; }
.contador { font-size: .8rem; color: #7a8a98; margin-bottom: 12px; font-weight: 500; animation: slideUp .45s ease-out .3s both; }
.toast { position: fixed; bottom: 28px; left: 50%; transform: translateX(-50%) translateY(12px); background: #1c2838; color: #f0f2f5; padding: 12px 24px; border-radius: 8px; font-weight: 500; font-size: .9rem; z-index: 999; opacity: 0; transition: all .25s ease; pointer-events: none; }
.toast.show { opacity: 1; transform: translateX(-50%) translateY(0); }
.footer { position: fixed; bottom: 12px; left: 0; right: 0; display: flex; justify-content: space-between; padding: 0 16px; font-size: .65rem; color: rgba(44, 62, 106, .3); pointer-events: none; z-index: 0; letter-spacing: .08em; text-transform: uppercase; }
.modal-overlay { position: fixed; inset: 0; background: rgba(10, 20, 30, .5); display: flex; align-items: center; justify-content: center; z-index: 1000; opacity: 0; transition: opacity .2s ease; pointer-events: none; }
.modal-overlay.show { opacity: 1; pointer-events: auto; }
.modal-card { background: #fff; border-radius: 10px; padding: 28px 32px; max-width: 440px; width: 92%; box-shadow: 0 12px 40px rgba(0,0,0,.15); text-align: center; transform: translateY(8px); transition: transform .25s ease; }
.modal-overlay.show .modal-card { transform: translateY(0); }
.modal-card p { font-size: .95rem; color: #1c2838; margin-bottom: 24px; line-height: 1.6; white-space: pre-wrap; }
.modal-card .botones { display: flex; gap: 12px; justify-content: center; }
.modal-card button { padding: 9px 22px; border-radius: 8px; border: none; font-weight: 500; font-size: .875rem; cursor: pointer; transition: all .15s ease; font-family: inherit; }
.modal-card .btn-si { background: #3a8a5a; color: #fff; }
.modal-card .btn-si:hover { background: #2a7a4a; }
.modal-card .btn-no { background: #f0f2f5; color: #1c2838; }
.modal-card .btn-no:hover { background: #e0e4e8; }
.modal-card .btn-ok { background: #2c5f8a; color: #fff; min-width: 110px; }
.modal-card .btn-ok:hover { background: #1e4a70; }
@keyframes slideUp { from { opacity: 0; transform: translateY(12px); } to { opacity: 1; transform: translateY(0); } }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
@keyframes rowIn { from { opacity: 0; transform: translateX(-14px); } to { opacity: 1; transform: translateX(0); } }
.estado.andando::before { content: ''; display: inline-block; width: 7px; height: 7px; border-radius: 50%; background: #3a8a5a; margin-right: 6px; }
.badge-novato, .badge-profesional, .badge-none { display: inline-block; padding: 2px 10px; border-radius: 4px; font-size: .75rem; font-weight: 500; }
.badge-novato { background: #eef2f6; color: #3a6a9a; }
.badge-profesional { background: #eef2f6; color: #2a7a48; }
.badge-none { background: transparent; color: #8a9ab0; font-weight: 400; }
.filtros-cat { display: flex; gap: 6px; flex-wrap: wrap; margin-bottom: 12px; animation: slideUp .45s ease-out .3s both; }
.filtro-cat { padding: 5px 14px; font-size: .8rem; border-radius: 6px; border: 1px solid #dce0e8; background: #f4f6f9; color: #1c2838; cursor: pointer; font-weight: 500; transition: all .15s ease; font-family: inherit; }
.filtro-cat:hover { background: #e8ecf0; }
.filtro-cat.activo { background: #1c2838; color: #fff; border-color: #1c2838; }
.filtro-cat.activo:hover { background: #0f1828; }
.modal-card.estadisticas { max-width: 620px; text-align: left; }
.est-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 18px 0; }
.est-card { background: #f4f6f9; border-radius: 8px; padding: 14px; text-align: center; border: 1px solid #eaecf0; }
.est-card .num { font-size: 1.6rem; font-weight: 700; color: #1c2838; display: block; }
.est-card .lab { font-size: .75rem; color: #7a8a98; margin-top: 3px; }
.est-card.verde-card .num { color: #2a8a48; }
.est-card.rojo-card .num { color: #b05046; }
.est-seccion { margin: 18px 0; }
.est-seccion h3 { font-size: .85rem; color: #1c2838; margin-bottom: 10px; font-weight: 600; }
.est-barra-wrap { display: flex; align-items: center; gap: 10px; margin-bottom: 8px; }
.est-barra-label { min-width: 110px; font-size: .8rem; color: #1c2838; font-weight: 500; }
.est-barra-track { flex: 1; height: 20px; background: #eaecf0; border-radius: 4px; overflow: hidden; }
.est-barra-fill { height: 100%; border-radius: 4px; transition: width .8s ease; background: #2c5f8a; }
.est-barra-fill.verde { background: #3a8a5a; }
.est-barra-fill.rojo { background: #b85048; }
.est-barra-num { min-width: 36px; font-size: .85rem; font-weight: 600; color: #1c2838; text-align: right; }
.est-tiempo { font-size: .8rem; color: #7a8a98; margin-left: 120px; margin-top: -2px; margin-bottom: 10px; }
.conexion { display: inline-flex; align-items: center; gap: 6px; }
.con-indicator { width: 8px; height: 8px; border-radius: 50%; display: inline-block; transition: background .3s ease; }
.con-indicator.online { background: #3a8a5a; box-shadow: 0 0 4px #3a8a5a; }
.con-indicator.offline { background: #b85048; box-shadow: 0 0 4px #b85048; }
.con-indicator.checking { background: #c9953e; box-shadow: 0 0 4px #c9953e; animation: pulse 1s infinite; }
@keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: .4; } }
@media (max-width: 700px) {
  .header-wrap { padding: 18px 20px 14px; }
  .header { gap: 12px; }
  .header h1 { font-size: 1.3rem; }
  .header img { height: 65px; }
  .main-content { padding: 18px 16px 70px; }
  .barra button, .barra input, .barra select { font-size: .8rem; padding: 9px 14px; }
  .barra input { min-width: 80px; }
  .seccion { padding: 16px 18px; }
  .buscar-wrap input { font-size: .85rem; }
  th, td { padding: 8px 10px; }
  .est-grid { grid-template-columns: 1fr; }
}
</style>
</head>
<body>
<div class="container" id="app">
  <div class="header-wrap">
    <div class="header">
      <img src="{{ logo_izq }}" alt="Logo">
      <h1>Burrotón San Benito José</h1>
      <img src="{{ logo_der }}" alt="Logo">
    </div>
  </div>
  <div class="main-content">
  <div class="seccion">
    <h2>Registro de participantes</h2>
    <div class="fila">
      <input id="dorsal-input" placeholder="Número" size="6">
      <input id="nombre-input" placeholder="Nombre del corredor">
      <select id="categoria-input">
        <option value="">Sin categoría</option>
        <option value="Novato">Novato</option>
        <option value="Profesional">Profesional</option>
      </select>
      <button onclick="registrar(this)">Registrar</button>
      <button onclick="document.getElementById('excel-input').click()">Importar Excel</button>
      <input type="file" id="excel-input" accept=".xlsx" style="display:none" onchange="importarExcel(this.files[0])">
    </div>
  </div>
  <div class="seccion" id="seccion-carrera">
    <h2>Control de carrera</h2>
    <div class="fila">
      <div id="estado-carrera" class="estado parada">Carrera no iniciada</div>
      <button id="btn-iniciar" class="verde" onclick="iniciar()">Iniciar carrera</button>
      <button id="btn-finalizar" class="rojo" onclick="finalizar()" style="display:none">Finalizar carrera</button>
      <button id="btn-limpiar" class="" onclick="limpiar()" style="display:none">Limpiar datos</button>
    </div>
    <div class="fila espaciada">
      <select id="modo-registro" onchange="cambioModo(this)">
        <option value="">Todos</option>
        <option value="Novato">Solo Novatos</option>
        <option value="Profesional">Solo Profesionales</option>
      </select>
      <input id="llegada-input" placeholder="Número dorsal">
      <button class="verde" onclick="llegada(this)">Registrar llegada</button>
    </div>
  </div>
  <div class="seccion">
    <h2>Reportes</h2>
    <div class="fila">
      <button onclick="resultados()">Ver resultados</button>
      <button onclick="reporte()">Descargar reporte</button>
      <button onclick="estadisticas()">Estadísticas</button>
    </div>
  </div>
  <div class="buscar-wrap">
    <input id="buscar-input" placeholder="Buscar por número o nombre..." oninput="cargar()">
  </div>
  <div class="filtros-cat" id="filtros-cat">
    <button class="filtro-cat activo" data-cat="" onclick="filtrarCategoria('')">Todos</button>
    <button class="filtro-cat" data-cat="Novato" onclick="filtrarCategoria('Novato')">Novato</button>
    <button class="filtro-cat" data-cat="Profesional" onclick="filtrarCategoria('Profesional')">Profesional</button>
    <button class="filtro-cat" data-cat="Sin categoría" onclick="filtrarCategoria('Sin categoría')">Sin categoría</button>
  </div>
  <div class="contador" id="contador"></div>
  <div class="tabla-wrap">
    <table>
      <thead><tr><th>Núm.</th><th>Nombre</th><th>Cat.</th><th>Llegada</th><th>Pos. Cat.</th><th>Acción</th></tr></thead>
      <tbody id="tabla"></tbody>
    </table>
  </div>
</div>
</div>
<div class="footer"><span>TURTLELITE</span><span class="conexion"><span class="con-indicator" id="con-indicator"></span><span id="con-text">Conectando...</span></span></div>
<div id="toast" class="toast"></div>
<div class="modal-overlay" id="modal">
  <div class="modal-card">
    <p id="modal-msg"></p>
    <div class="botones" id="modal-botones"></div>
  </div>
</div>
<script>
let filtroActual = '';
let enviando = false;
function esc(t) { return String(t).replace(/[&<>"']/g, function(m) { return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]; }); }

function filtrarCategoria(cat) {
  filtroActual = cat;
  document.querySelectorAll('.filtro-cat').forEach(b => b.classList.toggle('activo', b.dataset.cat === cat));
  cargar();
}
function cambioModo(el) {
  const val = el.value;
  const cat = val === "Novato" ? "Novato" : val === "Profesional" ? "Profesional" : "";
  filtrarCategoria(cat);
}

function cargar() {
  const q = document.getElementById('buscar-input').value.trim();
  const url = q ? '/api/buscar?q=' + encodeURIComponent(q) : '/api/datos';
  fetch(url).then(r=>{ if(!r.ok) throw new Error('Error de conexión'); return r.json(); }).then(d=>{
    const tbody = document.getElementById('tabla');
    tbody.innerHTML = '';
    let corredores = d.corredores||[];
    if (filtroActual) {
      corredores = corredores.filter(c => (c.categoria||'Sin categoría') === filtroActual);
    }
    corredores.forEach((c, i) => {
      const tr = document.createElement('tr');
      tr.style.animation = 'rowIn .35s ease-out both';
      tr.style.animationDelay = Math.min(i * 0.025, 0.6) + 's';
      const llegada = c.tiempo_llegada || '—';
      const posCat = c.posicion_categoria ? '#' + c.posicion_categoria : '—';
      const cat = c.categoria || '';
      const badge = cat ? '<span class="badge-' + cat.toLowerCase() + '">' + cat + '</span>' : '<span class="badge-none">—</span>';
      tr.innerHTML = '<td>' + esc(c.dorsal) + '</td><td>' + esc(c.nombre) + '</td><td>' + badge + '</td><td>' + llegada + '</td><td>' + esc(posCat) + '</td><td></td>';
      const btn = document.createElement('button');
      btn.className = 'rojo';
      btn.style.cssText = 'padding:4px 10px;font-size:.8rem';
      btn.textContent = '✕';
      btn.onclick = () => borrar(c.dorsal);
      tr.lastChild.appendChild(btn);
      tbody.appendChild(tr);
    });
    const total = d.corredores ? d.corredores.length : 0;
    const visibles = corredores.length;
    document.getElementById('contador').textContent = 'Total: ' + total + ' corredores' + (filtroActual ? ' — mostrando ' + visibles + ' ' + filtroActual : '');
    const est = document.getElementById('estado-carrera');
    if (d.carrera_iniciada) {
      est.textContent = 'Carrera en curso — Salida: ' + d.hora_inicio;
      est.className = 'estado andando';
      document.getElementById('btn-iniciar').disabled = true;
      document.getElementById('btn-finalizar').style.display = '';
      document.getElementById('btn-limpiar').style.display = 'none';
    } else {
      est.textContent = 'Carrera no iniciada';
      est.className = 'estado parada';
      document.getElementById('btn-iniciar').disabled = false;
      document.getElementById('btn-finalizar').style.display = 'none';
      document.getElementById('btn-limpiar').style.display = '';
    }
  }).catch(e => { if(e.message !== 'Error de conexión') mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'); });
}
function toast(msg) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.classList.add('show');
  setTimeout(() => el.classList.remove('show'), 3000);
}
function mostrarModal(msg) {
  const modal = document.getElementById('modal');
  document.getElementById('modal-msg').textContent = msg;
  document.getElementById('modal-botones').innerHTML = '<button class="btn-ok" onclick="cerrarModal()">Aceptar</button>';
  modal.classList.add('show');
}
function cerrarModal() {
  document.getElementById('modal').classList.remove('show');
  document.querySelector('.modal-card').classList.remove('estadisticas');
}
function confirmarModal(msg) {
  return new Promise(resolve => {
    const modal = document.getElementById('modal');
    document.getElementById('modal-msg').textContent = msg;
    document.getElementById('modal-botones').innerHTML =
      '<button class="btn-si" onclick="cerrarModal(); resolveConfirm(true)">Sí</button>' +
      '<button class="btn-no" onclick="cerrarModal(); resolveConfirm(false)">No</button>';
    modal.classList.add('show');
    window.resolveConfirm = resolve;
  });
}
async function _fetch(url, opts, btn, reintentos = 2) {
  for (let i = 0; i <= reintentos; i++) {
    if (btn) btn.disabled = true;
    try {
      const ctrl = new AbortController();
      const timer = setTimeout(() => ctrl.abort(), 20000);
      const r = await fetch(url, { ...opts, signal: ctrl.signal });
      clearTimeout(timer);
      if (!r.ok) {
        if (r.status >= 500 && i < reintentos) {
          btn && (btn.disabled = false);
          await new Promise(r => setTimeout(r, 2000));
          continue;
        }
        const msg = r.status === 404 ? 'Recurso no encontrado' : r.status === 409 ? 'Conflicto al guardar' : r.status === 503 ? 'Servicio no disponible' : 'Error de conexión';
        throw new Error(msg);
      }
      return await r.json();
    } catch (e) {
      clearTimeout(timer);
      if (i < reintentos && (e.name === 'AbortError' || e.name === 'TypeError')) {
        btn && (btn.disabled = false);
        if (i === 0) toast('Reintentando...');
        await new Promise(r => setTimeout(r, 3000));
        continue;
      }
      const msg = e.name === 'AbortError' || e.name === 'TypeError'
        ? 'No se pudo conectar. El servidor puede estar despertando, intenta de nuevo.'
        : e.message;
      mostrarModal(msg);
      throw e;
    } finally {
      if (btn) btn.disabled = false;
    }
  }
}
function registrar(btn) {
  const dorsal = document.getElementById('dorsal-input').value.trim();
  const nombre = document.getElementById('nombre-input').value.trim();
  const categoria = document.getElementById('categoria-input').value;
  if (!dorsal || !nombre) return mostrarModal('Completa número y nombre.');
  _fetch('/api/registrar', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({dorsal, nombre, categoria}) }, btn)
    .then(d => {
      if (d.duplicado) {
        confirmarModal('El número ' + dorsal + ' ya está registrado.\\n¿Reemplazarlo por ' + nombre + '?').then(r => {
          if (!r) return;
          _fetch('/api/registrar', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({dorsal, nombre, categoria, reemplazar: true}) }, btn)
            .then(d2 => { if(d2.error) mostrarModal(d2.error); else { document.getElementById('dorsal-input').value=''; document.getElementById('nombre-input').value=''; toast('Corredor reemplazado'); cargar(); } });
        });
      } else if (d.error) {
        mostrarModal(d.error);
      } else {
        document.getElementById('dorsal-input').value='';
        document.getElementById('nombre-input').value='';
        toast('Corredor registrado');
        cargar();
      }
    });
}
function iniciar() {
  const btn = document.getElementById('btn-iniciar');
  btn.disabled = true;
  _fetch('/api/iniciar', {method:'POST'}, btn).then(d=> { if(d.error) mostrarModal(d.error); else { toast('Carrera iniciada'); cargar(); } });
}
function llegada(btn) {
  if (enviando) return;
  const dorsal = document.getElementById('llegada-input').value.trim();
  if (!dorsal) return mostrarModal('Ingresa un número.');
  enviando = true;
  const categoria = document.getElementById('modo-registro').value;
  _fetch('/api/llegada', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({dorsal, categoria})}, btn)
    .then(d=> { if(d.error) mostrarModal(d.error); else { document.getElementById('llegada-input').value=''; cargar(); if(d.mensaje) toast(d.mensaje); } })
    .finally(() => { enviando = false; });
}
function finalizar() {
  confirmarModal('¿Finalizar la carrera?').then(r => {
    if (!r) return;
    fetch('/api/finalizar', {method:'POST'}).then(r=>r.json()).then(d=> { if(d.error) mostrarModal(d.error); else toast('Carrera finalizada'); cargar(); }).catch(e => mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'));
  });
}
function limpiar() {
  confirmarModal('¿Borrar todos los corredores y reiniciar la carrera?').then(r => {
    if (!r) return;
    fetch('/api/limpiar', {method:'POST'}).then(r=>r.json()).then(d=> { if(d.error) mostrarModal(d.error); else toast('Datos eliminados'); cargar(); }).catch(e => mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'));
  });
}
function resultados() {
  fetch('/api/resultados').then(r=>r.json()).then(d=> {
    if(d.error) return mostrarModal(d.error);
    const cats = {};
    d.llegados.forEach(c => {
      if (!cats[c.categoria]) cats[c.categoria] = [];
      cats[c.categoria].push(c);
    });
    let txt = 'RESULTADOS\\n' + '='.repeat(30) + '\\n';
    const ordenCats = ['Novato', 'Profesional', 'Sin categoría'];
    ordenCats.forEach(cat => {
      if (!cats[cat]) return;
      txt += '\\n  🏅 ' + cat.toUpperCase() + '\\n' + '-'.repeat(20) + '\\n';
      cats[cat].forEach(c => { txt += '  #' + c.pos_cat + '  ' + c.dorsal + '  ' + c.nombre + '  ' + c.tiempo + '\\n'; });
    });
    mostrarModal(txt);
  }).catch(e => mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'));
}
function reporte() {
  window.location.href = '/api/reporte';
}
function estadisticas() {
  fetch('/api/estadisticas').then(r=>r.json()).then(d => {
    if (d.error) return mostrarModal(d.error);
    const cats = ['Novato', 'Profesional', 'Sin categoría'];
    const maxCat = Math.max(1, ...cats.map(c => d.por_categoria[c] || 0));
    const maxArr = Math.max(1, ...cats.map(c => d.llegados_por_categoria[c] || 0));
    const pct = function(val, max) { return Math.round((val / max) * 100); };
    let html = '<div class="est-grid">';
    html += '<div class="est-card"><span class="num">' + d.total + '</span><span class="lab">Total corredores</span></div>';
    html += '<div class="est-card verde-card"><span class="num">' + d.llegados + '</span><span class="lab">Llegadas</span></div>';
    html += '<div class="est-card rojo-card"><span class="num">' + d.pendientes + '</span><span class="lab">Pendientes</span></div>';
    const catsReg = cats.filter(c => d.por_categoria[c]).length;
    html += '<div class="est-card"><span class="num">' + catsReg + '</span><span class="lab">Categorías</span></div>';
    html += '</div>';
    html += '<div class="est-seccion"><h3>Corredores por categoría</h3>';
    cats.forEach(c => {
      const val = d.por_categoria[c] || 0;
      const w = pct(val, maxCat);
      html += '<div class="est-barra-wrap"><span class="est-barra-label">' + c + '</span>';
      html += '<div class="est-barra-track"><div class="est-barra-fill" style="width:' + w + '%"></div></div>';
      html += '<span class="est-barra-num">' + val + '</span></div>';
    });
    html += '</div>';
    html += '<div class="est-seccion"><h3>Llegadas por categoría</h3>';
    cats.forEach(c => {
      const val = d.llegados_por_categoria[c] || 0;
      const w = pct(val, maxArr);
      html += '<div class="est-barra-wrap"><span class="est-barra-label">' + c + '</span>';
      html += '<div class="est-barra-track"><div class="est-barra-fill verde" style="width:' + w + '%"></div></div>';
      html += '<span class="est-barra-num">' + val + '</span></div>';
    });
    html += '</div>';
    if (Object.keys(d.tiempo_promedio).length) {
      html += '<div class="est-seccion"><h3>⏱ Tiempo promedio por categoría</h3>';
      cats.forEach(c => {
        if (d.tiempo_promedio[c]) {
          html += '<div class="est-tiempo"><strong>' + c + ':</strong> ' + d.tiempo_promedio[c] + '</div>';
        }
      });
      html += '</div>';
    }
    const modal = document.getElementById('modal');
    document.getElementById('modal-msg').innerHTML = html;
    document.getElementById('modal-botones').innerHTML = '<button class="btn-ok" onclick="cerrarModal()">Cerrar</button>';
    modal.classList.add('show');
    document.querySelector('.modal-card').classList.add('estadisticas');
  }).catch(e => mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'));
}
function borrar(dorsal) {
  confirmarModal('¿Borrar corredor número ' + dorsal + '?').then(r => {
    if (!r) return;
    fetch('/api/borrar', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({dorsal}) })
      .then(r=>r.json()).then(d=> { if(d.error) mostrarModal(d.error); else toast('Corredor eliminado'); cargar(); }).catch(e => mostrarModal('No se pudo conectar con el servidor. Verifica que el servidor esté encendido.'));
  });
}
function importarExcel(file) {
  if (!file) return;
  const form = new FormData();
  form.append('excel', file);
  const inp = document.getElementById('excel-input');
  fetch('/api/importar_excel', { method:'POST', body: form })
    .then(r => { console.log('Status:', r.status); return r.text().then(t => { console.log('Respuesta:', t); return JSON.parse(t); }); })
    .then(d => { if(d.error) mostrarModal(d.error); else toast(d.mensaje); cargar(); })
    .catch(e => { console.error('Error import:', e); mostrarModal('Error al importar: ' + e.message); })
    .finally(() => { inp.value = ''; });
}
function actualizarConexion(estado) {
  const dot = document.getElementById('con-indicator');
  const txt = document.getElementById('con-text');
  dot.className = 'con-indicator ' + estado;
  txt.textContent = estado === 'online' ? 'Conectado' : estado === 'offline' ? 'Desconectado' : 'Conectando...';
}
async function chequearConexion() {
  actualizarConexion('checking');
  try {
    const ctrl = new AbortController();
    const timer = setTimeout(() => ctrl.abort(), 8000);
    const r = await fetch('/api/datos', { signal: ctrl.signal });
    clearTimeout(timer);
    if (r.ok) actualizarConexion('online');
    else actualizarConexion('offline');
  } catch {
    actualizarConexion('offline');
  }
}
setInterval(chequearConexion, 15000);
chequearConexion();
cargar();
</script>
</body>
</html>"""

@app.route("/")
def index():
    return render_template_string(HTML, logo_izq=LOGO_IZQ, logo_der=LOGO_DER)

@app.route("/api/datos")
def api_datos():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT dorsal, nombre, categoria, tiempo_llegada, posicion, posicion_categoria FROM corredores ORDER BY id")
        rows = cur.fetchall()
        corredores = []
        for r in rows:
            corredores.append({"dorsal": r["dorsal"], "nombre": r["nombre"], "categoria": r["categoria"] or "", "tiempo_llegada": r["tiempo_llegada"].isoformat() if r["tiempo_llegada"] else None, "posicion": r["posicion"], "posicion_categoria": r["posicion_categoria"]})
        cur.execute("SELECT iniciada, hora_inicio FROM carrera WHERE id = 1")
        c = cur.fetchone()
        cur.close()
        if not c:
            return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])
        return jsonify(carrera_iniciada=c["iniciada"], hora_inicio=c["hora_inicio"].isoformat() if c["hora_inicio"] else None, corredores=corredores)
    except Exception as e:
        traceback.print_exc()
        return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])

@app.route("/api/buscar")
def api_buscar():
    init_db()
    q = request.args.get("q", "").strip()
    conn = get_db()
    if not conn:
        return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT dorsal, nombre, categoria, tiempo_llegada, posicion, posicion_categoria FROM corredores WHERE dorsal ILIKE %s OR nombre ILIKE %s ORDER BY id LIMIT 200", (f"%{q}%", f"%{q}%"))
        rows = cur.fetchall()
        corredores = []
        for r in rows:
            corredores.append({"dorsal": r["dorsal"], "nombre": r["nombre"], "categoria": r["categoria"] or "", "tiempo_llegada": r["tiempo_llegada"].isoformat() if r["tiempo_llegada"] else None, "posicion": r["posicion"], "posicion_categoria": r["posicion_categoria"]})
        cur.execute("SELECT iniciada, hora_inicio FROM carrera WHERE id = 1")
        c = cur.fetchone()
        cur.close()
        if not c:
            return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])
        return jsonify(carrera_iniciada=c["iniciada"], hora_inicio=c["hora_inicio"].isoformat() if c["hora_inicio"] else None, corredores=corredores)
    except Exception as e:
        traceback.print_exc()
        return jsonify(carrera_iniciada=False, hora_inicio=None, corredores=[])

@app.route("/api/registrar", methods=["POST"])
def api_registrar():
    init_db()
    dorsal = request.json.get("dorsal", "").strip()
    nombre = request.json.get("nombre", "").strip()
    categoria = request.json.get("categoria", "").strip()
    reemplazar = request.json.get("reemplazar", False)
    if not dorsal or not nombre:
        return jsonify(error="Completa todos los campos."), 400
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    try:
        cur = conn.cursor()
        if reemplazar:
            cur.execute("""
                INSERT INTO corredores (dorsal, nombre, categoria) VALUES (%s, %s, %s)
                ON CONFLICT (dorsal) DO UPDATE SET nombre = %s, categoria = %s,
                tiempo_llegada = NULL, posicion = NULL, posicion_categoria = NULL
            """, (dorsal, nombre, categoria, nombre, categoria))
        else:
            try:
                cur.execute("INSERT INTO corredores (dorsal, nombre, categoria) VALUES (%s, %s, %s)", (dorsal, nombre, categoria))
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                cur.close()
                return jsonify(error="Ya existe un corredor con ese número.", duplicado=True), 409
        conn.commit()
        cur.close()
        return jsonify(ok=True)
    except Exception as e:
        traceback.print_exc()
        return jsonify(error="Error al registrar"), 500

@app.route("/api/iniciar", methods=["POST"])
def api_iniciar():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    conn.autocommit = False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM carrera WHERE id = 1 FOR UPDATE")
        cur.execute("SELECT iniciada FROM carrera WHERE id = 1")
        if cur.fetchone()[0]:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error="La carrera ya está iniciada."), 400
        ahora = datetime.now()
        cur.execute("UPDATE carrera SET iniciada = TRUE, hora_inicio = %s WHERE id = 1", (ahora,))
        conn.commit()
        conn.autocommit = True
        cur.close()
        return jsonify(ok=True)
    except Exception as e:
        conn.rollback()
        conn.autocommit = True
        traceback.print_exc()
        return jsonify(error="Error al iniciar carrera"), 500

@app.route("/api/llegada", methods=["POST"])
def api_llegada():
    init_db()
    dorsal = request.json.get("dorsal", "").strip()
    categoria_filtro = request.json.get("categoria", "").strip()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    conn.autocommit = False
    try:
        cur = conn.cursor()
        cur.execute("SELECT iniciada FROM carrera WHERE id = 1")
        if not cur.fetchone()[0]:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error="La carrera no ha iniciado."), 400
        cur.execute("SELECT id, nombre, categoria, tiempo_llegada FROM corredores WHERE dorsal = %s", (dorsal,))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error="Número no encontrado."), 404
        if row[3]:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error=f"{row[1]} ya llegó."), 400
        if categoria_filtro and (row[2] or "") != categoria_filtro:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error=f"{row[1]} es {(row[2] or 'sin categoría')}, no {categoria_filtro}."), 400
        ahora = datetime.now()
        cat = row[2] or "Sin categoría"
        cur.execute("SELECT id FROM carrera WHERE id = 1 FOR UPDATE")
        cur.execute("SELECT COUNT(*) FROM corredores WHERE tiempo_llegada IS NOT NULL")
        posicion = cur.fetchone()[0] + 1
        cur.execute("SELECT COUNT(*) FROM corredores WHERE tiempo_llegada IS NOT NULL AND COALESCE(categoria, '') = COALESCE(%s, '')", (row[2],))
        posicion_categoria = cur.fetchone()[0] + 1
        cur.execute("UPDATE corredores SET tiempo_llegada = %s, posicion = %s, posicion_categoria = %s WHERE id = %s", (ahora, posicion, posicion_categoria, row[0]))
        cur.execute("SELECT hora_inicio FROM carrera WHERE id = 1 FOR UPDATE")
        inicio = cur.fetchone()[0]
        trans = abs(ahora - inicio)
        h, resto = divmod(int(trans.total_seconds()), 3600)
        m, s = divmod(resto, 60)
        conn.commit()
        conn.autocommit = True
        cur.close()
        return jsonify(ok=True, mensaje=f"¡{row[1]} llegó! #{posicion_categoria} en {cat} — {h}h {m}m {s}s")
    except Exception as e:
        conn.rollback()
        conn.autocommit = True
        traceback.print_exc()
        return jsonify(error="Error al registrar llegada"), 500

@app.route("/api/resultados")
def api_resultados():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    try:
        cur = conn.cursor()
        cur.execute("SELECT hora_inicio FROM carrera WHERE id = 1")
        inicio = cur.fetchone()[0]
        if not inicio:
            cur.close()
            return jsonify(error="La carrera no ha iniciado."), 400
        cur.execute("SELECT dorsal, nombre, categoria, posicion, posicion_categoria, tiempo_llegada FROM corredores WHERE tiempo_llegada IS NOT NULL ORDER BY COALESCE(categoria, ''), posicion_categoria, id")
        rows = cur.fetchall()
        if not rows:
            cur.close()
            return jsonify(error="Aún no hay llegadas."), 400
        res = []
        for r in rows:
            trans = abs(r[5] - inicio)
            h, resto = divmod(int(trans.total_seconds()), 3600)
            m, s = divmod(resto, 60)
            res.append({"pos": r[3], "pos_cat": r[4], "dorsal": r[0], "nombre": r[1], "categoria": r[2] or "Sin categoría", "tiempo": f"{h}h {m}m {s}s"})
        cur.close()
        return jsonify(llegados=res)
    except Exception as e:
        traceback.print_exc()
        return jsonify(error="Error al obtener resultados"), 500

@app.route("/api/finalizar", methods=["POST"])
def api_finalizar():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    conn.autocommit = False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM carrera WHERE id = 1 FOR UPDATE")
        cur.execute("SELECT iniciada FROM carrera WHERE id = 1")
        if not cur.fetchone()[0]:
            cur.close()
            conn.rollback()
            conn.autocommit = True
            return jsonify(error="La carrera no está iniciada."), 400
        cur.execute("UPDATE carrera SET iniciada = FALSE WHERE id = 1")
        conn.commit()
        conn.autocommit = True
        cur.close()
        return jsonify(ok=True)
    except Exception as e:
        conn.rollback()
        conn.autocommit = True
        traceback.print_exc()
        return jsonify(error="Error al finalizar carrera"), 500

@app.route("/api/importar_excel", methods=["POST"])
def api_importar_excel():
    init_db()
    if "excel" not in request.files:
        return jsonify(error="No se envió archivo."), 400
    file = request.files["excel"]
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        header_names = [str(h).strip() if h else "" for h in headers]
        col_dorsal = None
        col_nombre = None
        col_categoria = None
        for i, h in enumerate(header_names):
            h_lower = h.lower()
            if "dorsal" in h_lower or "numero" in h_lower or "num" in h_lower:
                col_dorsal = i
            if "nombre" in h_lower or "corredor" in h_lower or "apellido" in h_lower or "participante" in h_lower:
                col_nombre = i
            if "categoria" in h_lower or "categoría" in h_lower or "cat" in h_lower or "tipo" in h_lower or "nivel" in h_lower or "clase" in h_lower or "grupo" in h_lower or "division" in h_lower or "división" in h_lower or "novato" in h_lower or "profesional" in h_lower:
                col_categoria = i
        if col_dorsal is None or col_nombre is None:
            return jsonify(error="No se encontraron columnas 'Número' y 'Nombre'."), 400
        conn = get_db()
        if not conn:
            return jsonify(error="Base de datos no disponible"), 503
        cur = conn.cursor()
        cont = 0
        ejemplo_valor = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            dorsal = str(row[col_dorsal]).strip() if row[col_dorsal] is not None else ""
            nombre = str(row[col_nombre]).strip() if row[col_nombre] is not None else ""
            if not dorsal or not nombre:
                continue
            categoria = ""
            if col_categoria is not None and row[col_categoria] is not None:
                raw = row[col_categoria]
                ejemplo_valor = repr(raw)
                val = str(raw).strip().lower()
                if "novato" in val:
                    categoria = "Novato"
                elif "profesional" in val:
                    categoria = "Profesional"
                else:
                    ejemplo_valor = f"'{raw}' (no coincide)"
            cur.execute("""
                INSERT INTO corredores (dorsal, nombre, categoria) VALUES (%s, %s, %s)
                ON CONFLICT (dorsal) DO UPDATE SET nombre = %s, categoria = %s,
                tiempo_llegada = NULL, posicion = NULL, posicion_categoria = NULL
            """, (dorsal, nombre, categoria, nombre, categoria))
            conn.commit()
            cont += 1
        cur.close()
        info_cat = f" Columna: '{header_names[col_categoria] if col_categoria is not None else 'no detectada'}'."
        info_val = f" Valor ejemplo: {ejemplo_valor}." if ejemplo_valor else ""
        return jsonify(mensaje=f"Se importaron {cont} corredores.{info_cat}{info_val}")
    except Exception as e:
        traceback.print_exc()
        return jsonify(error="Error al procesar el Excel"), 500

@app.route("/api/estadisticas")
def api_estadisticas():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT COUNT(*) as total FROM corredores")
        total = cur.fetchone()["total"]
        cur.execute("SELECT COUNT(*) as total FROM corredores WHERE tiempo_llegada IS NOT NULL")
        llegados = cur.fetchone()["total"]
        pendientes = total - llegados
        cur.execute("SELECT CASE WHEN categoria = '' OR categoria IS NULL THEN 'Sin categoría' ELSE INITCAP(categoria) END as cat, COUNT(*) as cnt FROM corredores GROUP BY cat ORDER BY cat")
        rows = cur.fetchall()
        por_categoria = {}
        for r in rows:
            por_categoria[r["cat"]] = r["cnt"]
        cur.execute("SELECT CASE WHEN categoria = '' OR categoria IS NULL THEN 'Sin categoría' ELSE INITCAP(categoria) END as cat, COUNT(*) as cnt FROM corredores WHERE tiempo_llegada IS NOT NULL GROUP BY cat ORDER BY cat")
        rows = cur.fetchall()
        llegados_por_categoria = {}
        for r in rows:
            llegados_por_categoria[r["cat"]] = r["cnt"]
        cur.execute("SELECT hora_inicio FROM carrera WHERE id = 1")
        inicio_row = cur.fetchone()
        inicio = inicio_row["hora_inicio"] if inicio_row else None
        tiempo_promedio = {}
        if inicio and llegados > 0:
            cur.execute("SELECT CASE WHEN categoria = '' OR categoria IS NULL THEN 'Sin categoría' ELSE INITCAP(categoria) END as cat, AVG(EXTRACT(EPOCH FROM (tiempo_llegada - %s))) as avg_secs FROM corredores WHERE tiempo_llegada IS NOT NULL GROUP BY cat", (inicio,))
            for r in cur.fetchall():
                cat = r["cat"]
                secs = int(r["avg_secs"])
                h, resto = divmod(secs, 3600)
                m, s = divmod(resto, 60)
                tiempo_promedio[cat] = f"{h}h {m}m {s}s"
        cur.execute("SELECT iniciada FROM carrera WHERE id = 1")
        c = cur.fetchone()
        cur.close()
        return jsonify(
            total=total,
            por_categoria=por_categoria,
            llegados=llegados,
            pendientes=pendientes,
            llegados_por_categoria=llegados_por_categoria,
            tiempo_promedio=tiempo_promedio,
            carrera_iniciada=c["iniciada"] if c else False,
            hora_inicio=inicio.isoformat() if inicio else None
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify(error="Error al obtener estadísticas"), 500

@app.route("/api/limpiar", methods=["POST"])
def api_limpiar():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    conn.autocommit = False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM carrera WHERE id = 1 FOR UPDATE")
        cur.execute("DELETE FROM corredores")
        cur.execute("UPDATE carrera SET iniciada = FALSE, hora_inicio = NULL WHERE id = 1")
        conn.commit()
        conn.autocommit = True; cur.close()
        return jsonify(ok=True)
    except Exception as e:
        conn.rollback(); conn.autocommit = True
        traceback.print_exc()
        return jsonify(error="Error al limpiar datos"), 500

@app.route("/api/borrar", methods=["POST"])
def api_borrar():
    init_db()
    dorsal = request.json.get("dorsal", "").strip()
    if not dorsal:
        return jsonify(error="Número requerido."), 400
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    conn.autocommit = False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM carrera WHERE id = 1 FOR UPDATE")
        cur.execute("DELETE FROM corredores WHERE dorsal = %s", (dorsal,))
        cur.execute("SELECT id FROM corredores WHERE tiempo_llegada IS NOT NULL ORDER BY tiempo_llegada")
        for i, row in enumerate(cur.fetchall(), 1):
            cur.execute("UPDATE corredores SET posicion = %s WHERE id = %s", (i, row[0]))
        cur.execute("SELECT DISTINCT COALESCE(categoria, '') as cat FROM corredores WHERE tiempo_llegada IS NOT NULL")
        for cat_row in cur.fetchall():
            cur.execute("SELECT id FROM corredores WHERE tiempo_llegada IS NOT NULL AND COALESCE(categoria, '') = %s ORDER BY tiempo_llegada", (cat_row[0],))
            for i, row in enumerate(cur.fetchall(), 1):
                cur.execute("UPDATE corredores SET posicion_categoria = %s WHERE id = %s", (i, row[0]))
        conn.commit()
        conn.autocommit = True; cur.close()
        return jsonify(ok=True)
    except Exception as e:
        conn.rollback(); conn.autocommit = True
        traceback.print_exc()
        return jsonify(error="Error al borrar"), 500

@app.route("/api/reporte")
def api_reporte():
    init_db()
    conn = get_db()
    if not conn:
        return jsonify(error="Base de datos no disponible"), 503
    try:
        cur = conn.cursor()
        cur.execute("SELECT hora_inicio FROM carrera WHERE id = 1")
        inicio = cur.fetchone()[0]
        if not inicio:
            cur.close()
            return jsonify(error="La carrera no ha iniciado."), 400
        cur.execute("SELECT posicion, posicion_categoria, dorsal, nombre, categoria, tiempo_llegada FROM corredores WHERE tiempo_llegada IS NOT NULL ORDER BY COALESCE(categoria, ''), posicion_categoria, id")
        rows = cur.fetchall()
        cur.close()
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Burrotón"
        cats_order = ["Novato", "Profesional"]
        existing = sorted(set(r[4] or "Sin categoría" for r in rows))
        for c in existing:
            if c not in cats_order:
                cats_order.append(c)
        headers = ["Posición Gral", "Pos. Categoría", "Núm.", "Nombre", "Tiempo Llegada", "Tiempo Transcurrido"]
        primero = True
        for cat in cats_order:
            cat_rows = [r for r in rows if (r[4] or "Sin categoría") == cat]
            if not cat_rows:
                continue
            if not primero:
                ws.append([])
            primero = False
            ws.append([f"{cat.upper()}"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for r in cat_rows:
                trans = abs(r[5] - inicio)
                h, resto = divmod(int(trans.total_seconds()), 3600)
                m, s = divmod(resto, 60)
                ws.append([r[0], r[1], r[2], r[3], r[5].isoformat(), f"{h}h {m}m {s}s"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_maraton.xlsx")
    except Exception as e:
        traceback.print_exc()
        return jsonify(error="Error al generar reporte"), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
